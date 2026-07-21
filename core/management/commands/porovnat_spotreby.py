"""
Porovna nasi spotrebu (MeterReading / Meter.consumption_for) s referencni
tabulkou (Spotreby_2026_FM.xlsx / Spotreby_2026_NJ.xlsx, ulozene vedle
tohoto souboru) po jednotlivych meridlech a obdobich.

Sloupce tabulky SPOTREBY jsou cteny POZICNE (4 metadatove sloupce, pak
leden az prosinec v tomto poradi) - popisky mesicu v hlavicce nejsou
vzdy spolehlive (napr. v listu TEPLO je sloupec pro cerven omylem
oznacen "srpna").

Referencni tabulka u meridel s podruznym meridlem (viz sloupec Master
v Zasobnik_sluzeb.xlsx) vykazuje jen ZBYTKOVOU spotrebu (celkovy stav
minus soucet PRIMYCH podruznych meridel) - napr. E_A7 je master pro
E_A1..E_A6, takze referencni hodnota E_A7 = surova spotreba E_A7 minus
soucet surove spotreby E_A1..E_A6. Skript proto pred porovnanim odecte
od kazdeho meridla s podruznymi meridly jejich surovou spotrebu (bez
rekurze - staci primi potomci, protoze jejich "surova" hodnota uz v
sobe zahrnuje pripadne jejich vlastni potomky).

Pouziti:
  python manage.py porovnat_spotreby --site FM
  python manage.py porovnat_spotreby --site NJ
  python manage.py porovnat_spotreby --site FM --tolerance 0.5
"""
import os

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from core.models import Meter, Period, Site

SHEETS = ["ELEKTRO", "VODA", "TEPLO", "RADIATORY"]
XLSX_FILES = {
    "FM": "Spotreby_2026_FM.xlsx",
    "NJ": "Spotreby_2026_NJ.xlsx",
}
POOL_FILE = "Zasobnik_sluzeb.xlsx"

STOP_PREFIXES = ("Celkem", "Fakturov", "Fakturac", "Rozdíl", "CELKEM")


class Command(BaseCommand):
    help = "Porovná spotřebu měřidel v databázi s referenční tabulkou po jednotlivých měřidlech/obdobích."

    def add_arguments(self, parser):
        parser.add_argument("--site", type=str, required=True)
        parser.add_argument("--tolerance", type=float, default=0.1)

    def _load_children_map(self, site_arg):
        """Vrati {master_code: [slave_code, ...]} pro dany areal, dle Zasobnik_sluzeb.xlsx."""
        pool_path = os.path.join(os.path.dirname(__file__), POOL_FILE)
        if not os.path.exists(pool_path):
            return {}
        wb = openpyxl.load_workbook(pool_path, data_only=True)
        ws = wb["List1"]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        children = {}
        for row in rows[1:]:
            d = dict(zip(header, row))
            areal = str(d.get("Areal") or "").strip()
            if site_arg.upper() not in areal.upper():
                continue
            master = d.get("Master")
            code = d.get("OM")
            if not master or not code:
                continue
            children.setdefault(str(master).strip(), []).append(str(code).strip())
        return children

    def handle(self, *args, **options):
        site_arg = options["site"]
        site = Site.objects.filter(name__icontains=site_arg).first()
        if not site:
            raise CommandError(f"Areál '{site_arg}' nenalezen.")

        xlsx_name = XLSX_FILES.get(site_arg.upper())
        if not xlsx_name:
            raise CommandError(f"Pro areál '{site_arg}' není znám referenční soubor (očekávám FM nebo NJ).")

        xlsx_path = os.path.join(os.path.dirname(__file__), xlsx_name)
        if not os.path.exists(xlsx_path):
            raise CommandError(f"Soubor {xlsx_path} nenalezen.")

        children_map = self._load_children_map(site_arg)

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        # reference[code][month] = spotreba za dany kalendarni mesic (1-12)
        reference = {}
        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            mode = None
            for row in ws.iter_rows(values_only=True):
                first = str(row[0]).strip() if row[0] else ""
                if first == "STAVY":
                    mode = None
                    continue
                if first == "SPOTŘEBY":
                    mode = "spotreby"
                    continue
                if mode != "spotreby" or not first:
                    continue
                if first.startswith(STOP_PREFIXES):
                    mode = None
                    continue
                code = first
                for month in range(1, 13):
                    col_idx = 3 + month  # 4 metadatove sloupce (0..3), pak leden na indexu 4
                    if col_idx >= len(row):
                        continue
                    value = row[col_idx]
                    if value is None:
                        continue
                    try:
                        value = float(value)
                    except (TypeError, ValueError):
                        continue
                    reference.setdefault(code, {})[month] = value

        periods = Period.objects.order_by("year", "month")
        meters = Meter.objects.filter(site=site).order_by("meter_type", "code")
        meters_by_code = {m.code: m for m in meters if m.code}

        tolerance = options["tolerance"]
        checked = 0
        mismatches = 0
        missing_reference = set()
        missing_children = set()

        for meter in meters:
            code = meter.code
            if not code:
                continue
            ref_row = reference.get(code)
            child_codes = children_map.get(code, [])
            child_meters = []
            for child_code in child_codes:
                child_meter = meters_by_code.get(child_code)
                if child_meter:
                    child_meters.append(child_meter)
                else:
                    missing_children.add(f"{code}->{child_code}")

            for period in periods:
                if not meter.readings.filter(period=period).exists():
                    continue
                ours = meter.consumption_for(period)
                if ours is None:
                    continue
                for child_meter in child_meters:
                    child_value = child_meter.consumption_for(period)
                    if child_value is not None:
                        ours -= child_value
                if ref_row is None:
                    missing_reference.add(code)
                    continue
                ref_value = ref_row.get(period.month)
                if ref_value is None:
                    continue
                checked += 1
                diff = float(ours) - ref_value
                if abs(diff) > tolerance:
                    mismatches += 1
                    self.stdout.write(self.style.ERROR(
                        f"  {code} {period}: databáze(zbytek)={ours}  tabulka={ref_value}  rozdíl={diff:+.3f}"
                    ))

        self.stdout.write("")
        if missing_children:
            self.stdout.write(self.style.WARNING(
                f"Podružná měřidla ze Zásobníku nenalezená v databázi: {', '.join(sorted(missing_children))}"
            ))
        if missing_reference:
            self.stdout.write(self.style.WARNING(
                f"Měřidla bez referenčních dat v tabulce: {', '.join(sorted(missing_reference))}"
            ))
        self.stdout.write(self.style.SUCCESS(
            f"Zkontrolováno {checked} dvojic měřidlo/období, {mismatches} neshod (tolerance {tolerance})."
        ))
