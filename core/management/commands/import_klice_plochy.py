"""
Import klicu-ploch z Excel souboru Klice_Plochy.xlsx.
Pro kazdy radek:
1. Najde ClientCard podle PopisKarty
2. Najde Unit podle IDPLOCHY - pokud neexistuje, VYTVORI ji (napr. u
   arealu NJ neni samostatny "plochy.xlsx" se zakladni vymerou jako
   u FM, kazda IDPLOCHY tam patri prave jedne karte, takze se pouzije
   primo sloupec PronajatoM jako vymera nove Plochy).
3. Vytvori nebo aktualizuje CardUnit s vymerou a sazbou

Pouziti:
  python manage.py import_klice_plochy cesta/k/souboru.xlsx --site FM
"""
import openpyxl
from decimal import Decimal
from django.core.management.base import BaseCommand
from core.models import Site, Unit, ClientCard, CardUnit


class Command(BaseCommand):
    help = "Importuje klice-plochy (CardUnit) z Excel souboru"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--site", type=str, default="FM")
        parser.add_argument(
            "--create-units", action="store_true",
            help=(
                "Pokud Plocha (Unit) podle IDPLOCHY neexistuje, vytvorit ji rovnou "
                "z tohoto souboru (vymera = PronajatoM, ucel = Ucel, jednotka = "
                "Jednotka). Pouzij jen kdyz kazda IDPLOCHY v souboru patri prave "
                "jedne karte (jinak by se vymera prepsala podle posledniho radku)."
            ),
        )

    def handle(self, *args, **options):
        site = Site.objects.filter(name__icontains=options["site"]).first()
        if not site:
            self.stdout.write(self.style.ERROR(f"Areál '{options['site']}' nenalezen."))
            return

        wb = openpyxl.load_workbook(options["xlsx_path"], read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        created = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            card_desc = str(data.get("PopisKarty") or "").strip()
            unit_code = str(data.get("IDPLOCHY") or "").strip()
            ucel = str(data.get("Ucel") or "").strip()
            jednotka = str(data.get("Jednotka") or "").strip() or "m2"
            pronajato = data.get("PronajatoM")
            sazba = data.get("SazbaKcMRok")
            aktivni = str(data.get("Aktivni") or "").strip().lower()

            if not card_desc or not unit_code or aktivni != "zapnuto":
                skipped += 1
                continue

            card = ClientCard.objects.filter(description=card_desc, is_active=True).first()
            if not card:
                self.stdout.write(f"  Karta nenalezena: {card_desc}")
                skipped += 1
                continue

            unit = Unit.objects.filter(code=unit_code, site=site).first()
            if not unit:
                # Zkusime hledat bez arealu (pro plochy bez site)
                unit = Unit.objects.filter(code=unit_code).first()
            if not unit and options["create_units"]:
                unit = Unit.objects.create(
                    site=site,
                    code=unit_code,
                    name=unit_code,
                    purpose=ucel,
                    area_m2=Decimal(str(pronajato)) if pronajato else None,
                    unit_type=jednotka,
                )
                self.stdout.write(f"  ++ Plocha vytvořena: {unit_code} ({pronajato} {jednotka})")
            if not unit:
                self.stdout.write(f"  Plocha nenalezena: {unit_code}")
                skipped += 1
                continue

            area_override = Decimal(str(pronajato)) if pronajato else None
            rate = Decimal(str(sazba)) if sazba and float(sazba) > 0 else None

            card_unit, was_created = CardUnit.objects.update_or_create(
                card=card,
                unit=unit,
                defaults={
                    "area_m2_override": area_override,
                    "rate_per_m2": rate,
                },
            )

            if was_created:
                created += 1
                self.stdout.write(f"  + {card_desc} / {unit_code}: {area_override} m² @ {rate} Kč")
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nHotovo: {created} vytvořeno, {updated} aktualizováno, {skipped} přeskočeno."
        ))
