"""
Import zasobniku sluzeb (mericí + polozky zasobniku) z Excelu.

Sloupce v Excelu: OM, Popis, Areal, Aktivni, Podruzny, Master,
typ_vypoctu, IMP_Predvyplnit, IMP_Spotreba, Cena_Jednotku, Pobocka,
IMP_Jednotka, Koef1, class, Virtual

Postup:
1) Pro tridy ELEKTRO/VODA/TEPLO vytvori/aktualizuje Meter pro KAZDY
   radek (vcetne podruznych) - vcetne hierarchie (Master -> parent_meter)
   a priznaku is_virtual.
2) Pro KORENOVE mericí (Master prazdny) v energetickych tridach
   vytvori odpovidajici ServicePoolItem (propojeny na dane mericí).
3) Pro tridu OSTATNI (vzdy korenova, bez mericí) vytvori ServicePoolItem
   bez vazby na mericí.

Pouziti: python manage.py import_zasobnik_sluzeb cesta/k/souboru.xlsx --site FM
"""
import openpyxl
from django.core.management.base import BaseCommand
from core.models import Site, Meter, ServicePoolItem


CLASS_TO_METER_TYPE = {
    "ELEKTRO": "electricity",
    "VODA": "water",
    "TEPLO": "heat",
}

CLASS_TO_INVOICE_CLASS = {
    "ELEKTRO": "electricity",
    "VODA": "water",
    "TEPLO": "heat",
    "OSTATNÍ": "other",
}

TYP_VYPOCTU_TO_ALLOCATION = {
    "K_CELKU": "percent",
    "K_PLOSE": "area_ratio",
    "PEVNA_KC": "fixed_amount",
}


class Command(BaseCommand):
    help = "Importuje zasobnik sluzeb (mericí + polozky zasobniku) z Excelu pro dany areal"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Cesta k Excel souboru")
        parser.add_argument("--site", type=str, required=True, help="Nazev arealu (napr. FM)")

    def handle(self, *args, **options):
        site = Site.objects.filter(name__icontains=options["site"]).first()
        if not site:
            self.stdout.write(self.style.ERROR(
                f"Areál '{options['site']}' nenalezen. Dostupné: "
                f"{list(Site.objects.values_list('name', flat=True))}"
            ))
            return

        wb = openpyxl.load_workbook(options["xlsx_path"], read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

        # --- Krok 1: vytvoreni/aktualizace mericí (bez parent_meter) ---
        self.stdout.write(f"--- Měřidla (areál {site}) ---")
        meters_by_code = {}
        meters_created = 0

        for row in rows:
            cls = row.get("class")
            if cls not in CLASS_TO_METER_TYPE:
                continue  # OSTATNI nema sve mericí

            code = str(row.get("OM") or "").strip()
            if not code:
                continue

            meter, created = Meter.objects.get_or_create(
                site=site,
                code=code,
                defaults={
                    "name": str(row.get("Popis") or code).strip(),
                    "meter_type": CLASS_TO_METER_TYPE[cls],
                    "unit_of_measure": str(row.get("IMP_Jednotka") or "").strip() or "kWh",
                    "is_virtual": bool(row.get("Virtual")),
                },
            )
            if not created:
                # aktualizace zakladnich udaju, kdyby se import spoustel znovu
                meter.name = str(row.get("Popis") or code).strip()
                meter.meter_type = CLASS_TO_METER_TYPE[cls]
                meter.is_virtual = bool(row.get("Virtual"))
                meter.save()
            else:
                meters_created += 1

            meters_by_code[code] = meter

        self.stdout.write(f"Měřidel vytvořeno: {meters_created} (celkem zpracováno {len(meters_by_code)})")

        # --- Krok 2: propojeni hierarchie (Master -> parent_meter) ---
        self.stdout.write("--- Propojuji hierarchii (Master) ---")
        linked = 0
        for row in rows:
            cls = row.get("class")
            if cls not in CLASS_TO_METER_TYPE:
                continue
            code = str(row.get("OM") or "").strip()
            master_code = str(row.get("Master") or "").strip()
            if not code or not master_code:
                continue

            meter = meters_by_code.get(code)
            parent = meters_by_code.get(master_code)
            if not meter or not parent:
                self.stdout.write(f"  Nelze propojit: {code} -> {master_code} (chybí měřidlo)")
                continue
            if meter.parent_meter_id != parent.id:
                meter.parent_meter = parent
                meter.save()
                linked += 1

        self.stdout.write(f"Propojeno hierarchií: {linked}")

        # --- Krok 3: ServicePoolItem pro korenova mericí (energie) ---
        self.stdout.write("--- Položky zásobníku (kořenová měřidla) ---")
        items_created = 0
        for row in rows:
            cls = row.get("class")
            if cls not in CLASS_TO_METER_TYPE:
                continue
            master_code = str(row.get("Master") or "").strip()
            if master_code:
                continue  # neni korenove

            code = str(row.get("OM") or "").strip()
            meter = meters_by_code.get(code)
            if not meter:
                continue

            allocation = TYP_VYPOCTU_TO_ALLOCATION.get(str(row.get("typ_vypoctu") or "").strip(), "")

            item, created = ServicePoolItem.objects.get_or_create(
                site=site,
                meter=meter,
                defaults={
                    "name": str(row.get("Popis") or code).strip(),
                    "invoice_class": CLASS_TO_INVOICE_CLASS[cls],
                    "default_allocation_type": allocation,
                },
            )
            if created:
                items_created += 1
                self.stdout.write(f"  + {item.name} ({CLASS_TO_INVOICE_CLASS[cls]})")

        # --- Krok 4: ServicePoolItem pro OSTATNI (bez mericí) ---
        self.stdout.write("--- Položky zásobníku (ostatní služby) ---")
        for row in rows:
            if row.get("class") != "OSTATNÍ":
                continue

            name = str(row.get("Popis") or row.get("OM") or "").strip()
            allocation = TYP_VYPOCTU_TO_ALLOCATION.get(str(row.get("typ_vypoctu") or "").strip(), "")

            item, created = ServicePoolItem.objects.get_or_create(
                site=site,
                name=name,
                meter=None,
                defaults={
                    "invoice_class": "other",
                    "default_allocation_type": allocation,
                },
            )
            if created:
                items_created += 1
                self.stdout.write(f"  + {item.name} (other)")

        self.stdout.write(self.style.SUCCESS(
            f"\nHotovo. Měřidel zpracováno: {len(meters_by_code)}, "
            f"položek zásobníku vytvořeno: {items_created}."
        ))
