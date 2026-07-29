"""
Diagnostika (jen cteni): pro dany seznam kodu meridel ukaze - co ma DB
nastavene jako Meter.parent_meter (a jeho children), surovy odecet,
"vlastni" (po odectu primych deti - stejna logika jako
billing/engine.py _owned_consumption / porovnat_spotreby.py), a hodnotu
z referencni tabulky Spotreby_2026_NJ.xlsx/Spotreby_2026_FM.xlsx pro
zadane obdobi - vse pohromade, aby se dalo primo videt, kde presne se
nase cislo rozchazi s referenci.

Pouziti:
  python manage.py diag_meridlo_hierarchie --site NJ --period=06/2026 --kody E_AB1_08,E_AB1_10,E_AB1_11,E_AB1_12
"""
import os

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from core.models import Meter, Period, Site

SHEETS = ["ELEKTRO", "VODA", "TEPLO", "RADIATORY"]
XLSX_FILES = {"FM": "Spotreby_2026_FM.xlsx", "NJ": "Spotreby_2026_NJ.xlsx"}
POOL_FILE = "Zasobnik_sluzeb.xlsx"
STOP_PREFIXES = ("Celkem", "Fakturov", "Fakturac", "Rozdíl", "CELKEM")


class Command(BaseCommand):
    help = "Detailní rozpad hierarchie + referenční hodnoty pro konkrétní seznam kódů měřidel."

    def add_arguments(self, parser):
        parser.add_argument("--site", required=True)
        parser.add_argument("--period", required=True, help="MM/RRRR")
        parser.add_argument("--kody", required=True, help="čárkou oddělené kódy měřidel")

    def _load_reference(self, xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        reference = {}
        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            mode = None
            for row in ws.iter_rows(values_only=True):
                first = str(row[0]).strip() if row[0] else ""
                if first == "STAVY":
                    mode = None
                    continue
                if first == "SPOTŘEBY":
                    mode = "spotreby"
                    continue
                if mode != "spotreby" or not first:
                    continue
                if first.startswith(STOP_PREFIXES):
                    mode = None
                    continue
                for month in range(1, 13):
                    col_idx = 3 + month
                    if col_idx >= len(row):
                        continue
                    value = row[col_idx]
                    if value is None:
                        continue
                    try:
                        value = float(value)
                    except (TypeError, ValueError):
                        continue
                    reference.setdefault(first, {})[month] = value
        return reference

    def handle(self, *args, **options):
        site_arg = options["site"]
        site = Site.objects.filter(name__icontains=site_arg).first()
        if not site:
            raise CommandError(f"Areál '{site_arg}' nenalezen.")

        month_str, year_str = options["period"].split("/")
        period = Period.objects.filter(year=int(year_str), month=int(month_str)).first()
        if period is None:
            raise CommandError(f"Období {options['period']} nenalezeno.")

        xlsx_name = XLSX_FILES.get(site_arg.upper())
        xlsx_path = os.path.join(os.path.dirname(__file__), xlsx_name) if xlsx_name else None
        reference = self._load_reference(xlsx_path) if xlsx_path and os.path.exists(xlsx_path) else {}

        codes = [c.strip() for c in options["kody"].split(",") if c.strip()]

        for code in codes:
            meter = Meter.objects.filter(site=site, code=code).first()
            self.stdout.write(self.style.WARNING(f"\n=== {code} ==="))
            if meter is None:
                self.stdout.write(self.style.ERROR("  Měřidlo nenalezeno v DB."))
                continue

            self.stdout.write(
                f"  DB parent_meter (Master) = {meter.parent_meter or '(žádný - kořenové měřidlo)'}"
            )
            children = list(meter.children.all())
            self.stdout.write(
                f"  DB children (Slave)      = {[c.code for c in children] or '(žádné)'}"
            )

            raw = meter.consumption_for(period)
            self.stdout.write(f"  surový odečet za {period} = {raw}")

            owned = raw
            if raw is not None:
                for child in children:
                    child_consumption = child.consumption_for(period)
                    if child_consumption is not None:
                        owned -= child_consumption
                        self.stdout.write(f"    - dítě {child.code}: {child_consumption}")
            self.stdout.write(f"  vlastní (po odečtu dětí) = {owned}")

            ref_row = reference.get(code)
            ref_value = ref_row.get(period.month) if ref_row else None
            self.stdout.write(f"  referenční tabulka ({xlsx_name}) = {ref_value if ref_value is not None else '(chybí)'}")

            if owned is not None and ref_value is not None:
                diff = float(owned) - ref_value
                self.stdout.write(
                    (self.style.SUCCESS if abs(diff) <= 0.1 else self.style.ERROR)(
                        f"  rozdíl (naše vlastní - reference) = {diff:+.3f}"
                    )
                )
