"""
Import klicu elektrina z Excel souboru Klice_Elektro.xlsx.
Pro kazdy radek vytvori AllocationKey na prislusne karte klienta.

Mapovani TYP_Polozky:
  K_CELKU -> vaha (hodnota = sloupec Jednotek, NE Podil - Podil je v
             puvodnim Excelu jen dopocitany napovedny udaj = Jednotek /
             soucet Jednotek pres vsechny karty se stejnym EL_KodOM;
             pouzitim primo Jednotek si system podily dopocitava sam
             a nezastarava pri zmene poctu najemcu). Typ klice je
             weighted_count, nebo person_count pokud EL_KodOM
             odpovida odberu na TUV (teplá užitková voda).
  K_PLOSE -> fixed_amount (hodnota = Mplochy * KCzaM / 12 = mesicni pausal)

EL_KodOM je kod meridla - hleda se v Meter.code pro dany areal.
ServicePoolItem se hleda podle meter. U K_CELKU se meridlo (slave kod,
napr. E_A1) uklada i do AllocationKey.meter - ne jako "podruzne meridlo"
ve smyslu vypoctu, ale aby jedna karta mohla mit vice K_CELKU radku
(ruzne kategorie/mistnosti) pro stejnou polozku Zasobniku soucasne.

Pouziti:
  python manage.py import_klice_elektro cesta/k/souboru.xlsx --site FM
"""
import openpyxl
from decimal import Decimal, ROUND_HALF_UP
from django.core.management.base import BaseCommand
from core.models import Site, Meter, ServicePoolItem, ClientCard, AllocationKey


TYP_TO_ALLOCATION = {
    "K_CELKU": "weighted_count",
    "K_PLOSE": "fixed_amount",
}


class Command(BaseCommand):
    help = "Importuje klice elektrina (AllocationKey) z Excel souboru"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--site", type=str, default="FM")

    def handle(self, *args, **options):
        site = Site.objects.filter(name__icontains=options["site"]).first()
        if not site:
            self.stdout.write(self.style.ERROR(f"Areál '{options['site']}' nenalezen."))
            return

        wb = openpyxl.load_workbook(options["xlsx_path"], read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        created = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            card_desc = str(data.get("PopisKarty") or "").strip()
            aktivni = str(data.get("Aktivni") or "").strip().lower()
            meter_code = str(data.get("EL_KodOM") or "").strip()
            typ = str(data.get("TYP_Polozky") or "").strip()
            jednotek = data.get("Jednotek")
            mplochy = data.get("Mplochy")
            kczam = data.get("KCzaM")

            if not card_desc or not meter_code or aktivni != "zapnuto":
                skipped += 1
                continue

            allocation_type = TYP_TO_ALLOCATION.get(typ)
            if not allocation_type:
                self.stdout.write(f"  Neznámý typ: {typ} ({card_desc} / {meter_code})")
                skipped += 1
                continue

            card = ClientCard.objects.filter(description=card_desc, is_active=True).first()
            if not card:
                self.stdout.write(f"  Karta nenalezena: {card_desc}")
                skipped += 1
                continue

            meter = Meter.objects.filter(code=meter_code, site=site).first()
            if not meter:
                self.stdout.write(f"  Měřidlo nenalezeno: {meter_code}")
                skipped += 1
                continue

            # Najdeme ServicePoolItem pro toto meridlo
            service_item = ServicePoolItem.objects.filter(meter=meter, site=site).first()
            if not service_item:
                # Zkusime najit hlavni polozku (bez parent - root meridlo)
                root = meter
                while root.parent_meter:
                    root = root.parent_meter
                service_item = ServicePoolItem.objects.filter(meter=root, site=site).first()

            if not service_item:
                self.stdout.write(f"  ServicePoolItem nenalezen pro měřidlo: {meter_code}")
                skipped += 1
                continue

            # Vypocet hodnoty klice
            if typ == "K_CELKU":
                jednotek_val = Decimal(str(jednotek)) if jednotek not in (None, "") else None
                if not jednotek_val:
                    self.stdout.write(f"  Přeskočeno (Jednotek=0/chybí): {card_desc} / {meter_code}")
                    skipped += 1
                    continue
                value = jednotek_val.quantize(Decimal("0.0001"))
                allocation_type = "person_count" if "TUV" in meter_code.upper() else "weighted_count"
            elif allocation_type == "fixed_amount":
                if mplochy and kczam and float(kczam) > 0:
                    # mesicni pausal = plocha * cena/m2/rok / 12
                    value = (Decimal(str(mplochy)) * Decimal(str(kczam)) / 12).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                else:
                    value = None
            else:
                value = None

            key, was_created = AllocationKey.objects.update_or_create(
                client_card=card,
                service_item=service_item,
                meter=meter if typ == "K_CELKU" else None,
                defaults={
                    "allocation_type": allocation_type,
                    "value": value,
                },
            )

            if was_created:
                created += 1
                self.stdout.write(f"  + {card_desc} / {meter_code} ({allocation_type}): {value}")
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nHotovo: {created} vytvořeno, {updated} aktualizováno, {skipped} přeskočeno."
        ))
