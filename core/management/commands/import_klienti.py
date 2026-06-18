import openpyxl
from django.core.management.base import BaseCommand
from core.models import Client


class Command(BaseCommand):
    help = "Importuje klienty z Excel souboru"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Cesta k Excel souboru")

    def handle(self, *args, **options):
        path = options["xlsx_path"]
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        created = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            if not data.get("NazevJmeno"):
                continue

            code = str(data.get("IDFIRMA") or "").strip()
            if Client.objects.filter(code=code).exists():
                skipped += 1
                continue

            Client.objects.create(
                code=code,
                name=str(data.get("NazevJmeno") or "").strip(),
                is_active=str(data.get("Aktivni") or "").strip().lower() == "ano",
                is_landlord=str(data.get("Pronajimatel") or "").strip().lower() == "ano",
                street=str(data.get("SidloUlice1") or "").strip(),
                street_number=str(data.get("SidloCislo1") or "").strip(),
                zip_code=str(data.get("SidloPSC1") or "").strip(),
                city=str(data.get("SidloMesto1") or "").strip(),
                ico=str(data.get("ICO") or "").strip(),
                dic=str(data.get("DIC") or "").strip(),
                vat_payer=str(data.get("PlatceDPH") or "").strip().lower() == "ano",
                bank_name=str(data.get("Banka") or "").strip(),
                bank_account=str(data.get("CUctu") or "").strip(),
                bank_code=str(data.get("KodBanky") or "").strip(),
                contact_email=str(data.get("Email") or "").strip(),
                contact_phone=str(data.get("Telefon") or "").strip(),
            )
            created += 1
            self.stdout.write(f"  Vytvořen: {data.get('NazevJmeno')}")

        self.stdout.write(self.style.SUCCESS(
            f"Hotovo: {created} klientů vytvořeno, {skipped} přeskočeno."
        ))
