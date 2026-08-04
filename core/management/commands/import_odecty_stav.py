"""
Import stavu meridel z jednoducheho plocheho Excelu (jeden list, sloupce
"<KATEGORIE>_KodOM"/"Stav"/"Jednotka"/"DatumOdectu", jeden odectovy den
pro vsechny radky) - jiny format nez puvodni import_odecty.py (psany
pro Spotreby_2026_FM.xlsx s vice mesici ve sloupcich najednou).

Stav k 1. dni mesice = odecet UZAVIRAJICI predchozi mesic (stejna
konvence jako import_odecty.py) - napr. stav k 1.8.2026 se ulozi jako
MeterReading pro obdobi 07/2026.

Sloupce se hledaji podle nazvu v hlavicce (ne pevne pozice), takze
prezije mirne odlisne nazvy napr. "VODA_KodOM" misto "ELEKTRO_KodOM".

Pouziti (staci holy nazev souboru, pokud lezi vedle tohoto prikazu -
viz _resolve_xlsx_path - Railway shell nemusi mit cwd v core/management/
commands/, takze relativni cesta z aktualniho adresare casto neexistuje):
  python manage.py import_odecty_stav Odecty_NJ_0108.xlsx --site NJ
"""
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from core.models import Meter, MeterReading, Period, Site


def _resolve_xlsx_path(xlsx_path):
    """Pokud zadana cesta neexistuje jak je (typicky holy nazev souboru
    spusteny z jineho cwd nez core/management/commands/), zkusi ji najit
    vedle tohoto command souboru - tam se zdrojove xlsx v tomhle repu
    tradicne verzuji (viz klice_NJ.xlsx, Spotreby_2026_NJ.xlsx apod.)."""
    path = Path(xlsx_path)
    if path.exists():
        return path
    fallback = Path(__file__).resolve().parent / path.name
    if fallback.exists():
        return fallback
    raise CommandError(f"Soubor {xlsx_path!r} nenalezen (ani vedle tohoto příkazu: {fallback}).")


class Command(BaseCommand):
    help = "Importuje stavy měřidel z jednoduchého plochého Excelu (kód/stav/datum, jeden odečtový den)."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--site", type=str, required=True)
        parser.add_argument("--sheet", type=str, default=None, help="Název listu (výchozí: první list).")

    def handle(self, *args, **options):
        site = Site.objects.filter(name__icontains=options["site"]).first()
        if not site:
            raise CommandError(f"Areál obsahující {options['site']!r} nenalezen.")

        xlsx_path = _resolve_xlsx_path(options["xlsx_path"])
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb[options["sheet"]] if options["sheet"] else wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("List je prázdný.")

        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_code = next((i for i, h in enumerate(headers) if h.endswith("_KodOM") or h == "KodOM"), None)
        col_stav = next((i for i, h in enumerate(headers) if h == "Stav"), None)
        col_datum = next((i for i, h in enumerate(headers) if h == "DatumOdectu"), None)

        missing = [
            name for name, col in [("<...>_KodOM", col_code), ("Stav", col_stav), ("DatumOdectu", col_datum)]
            if col is None
        ]
        if missing:
            raise CommandError(f"Nenalezeny sloupce {missing} v hlavičce: {headers}")

        created = updated = skipped = 0

        for row in rows[1:]:
            code = str(row[col_code]).strip() if row[col_code] else ""
            if not code:
                continue

            stav = row[col_stav]
            if stav is None:
                continue

            reading_date = row[col_datum]
            if reading_date is None:
                self.stdout.write(self.style.WARNING(f"  {code}: chybí DatumOdectu, přeskočeno"))
                skipped += 1
                continue
            if hasattr(reading_date, "date"):
                reading_date = reading_date.date()

            meter = Meter.objects.filter(site=site, code=code).first()
            if meter is None:
                self.stdout.write(self.style.WARNING(f"  {code}: měřidlo nenalezeno v areálu {site}"))
                skipped += 1
                continue

            # Stav k 1. dni mesice uzavira predchozi mesic (stejna konvence jako import_odecty.py).
            if reading_date.month == 1:
                period_year, period_month = reading_date.year - 1, 12
            else:
                period_year, period_month = reading_date.year, reading_date.month - 1

            period, _ = Period.objects.get_or_create(
                year=period_year, month=period_month, defaults={"status": "open"},
            )

            reading, was_created = MeterReading.objects.update_or_create(
                meter=meter, period=period,
                defaults={"value": stav, "reading_date": reading_date},
            )
            if was_created:
                created += 1
            else:
                updated += 1
            self.stdout.write(
                f"  {code}: stav={stav} k {reading_date} -> období {period} "
                f"({'nový' if was_created else 'aktualizován'})"
            )

        self.stdout.write(self.style.SUCCESS(f"\nVytvořeno {created}, aktualizováno {updated}, přeskočeno {skipped}."))
