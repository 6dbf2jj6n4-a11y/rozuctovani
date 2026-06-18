import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand
from core.models import Client, ClientCard, Site, Unit, CardUnit


def excel_date(value):
    """Převod Excel čísla na datum."""
    if isinstance(value, (int, float)):
        return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(value) - 2).date()
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    return None


class Command(BaseCommand):
    help = "Importuje karty klientů a plochy z Excel souborů"

    def add_arguments(self, parser):
        parser.add_argument("karty_xlsx", type=str, help="Cesta k souboru karty.xlsx")
        parser.add_argument("plochy_xlsx", type=str, help="Cesta k souboru plochy.xlsx")
        parser.add_argument("--site", type=str, default="FM", help="Název areálu (výchozí: FM)")

    def handle(self, *args, **options):
        site_name = options["site"]
        site = Site.objects.filter(name__icontains=site_name).first()
        if not site:
            self.stdout.write(self.style.ERROR(f"Areál '{site_name}' nenalezen. Dostupné areály: {list(Site.objects.values_list('name', flat=True))}"))
            return

        self.stdout.write(f"Používám areál: {site}")

        # 1. Import karet
        self.stdout.write("\n--- Import karet klientů ---")
        wb = openpyxl.load_workbook(options["karty_xlsx"], read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        cards_created = 0
        cards_map = {}  # IDK -> ClientCard instance

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if not data.get("IDK"):
                continue

            idk = int(data["IDK"])
            idfirma = str(data.get("IDFIRMA") or "").strip()
            client = Client.objects.filter(code=idfirma).first()

            if not client:
                self.stdout.write(f"  Firma nenalezena: {idfirma} (karta {idk})")
                continue

            if ClientCard.objects.filter(external_id=idk).exists():
                card = ClientCard.objects.get(external_id=idk)
                cards_map[idk] = card
                continue

            valid_from = excel_date(data.get("DatumOd")) or date.today()
            is_active = str(data.get("Aktivni") or "").strip().lower() == "zapnuto"

            card = ClientCard.objects.create(
                client=client,
                unit=None,
                external_id=idk,
                description=str(data.get("PopisKarty") or "").strip(),
                valid_from=valid_from,
                valid_to=None if is_active else valid_from,
                note=str(data.get("Poznamka") or "").strip(),
            )
            cards_map[idk] = card
            cards_created += 1
            self.stdout.write(f"  Vytvořena karta: {data.get('PopisKarty')}")

        self.stdout.write(f"Karet vytvořeno: {cards_created}")

        # 2. Import ploch
       # 2. Import ploch a propojení s kartami
        self.stdout.write("\n--- Import ploch a propojení s kartami ---")
        wb2 = openpyxl.load_workbook(options["plochy_xlsx"], read_only=True)
        ws2 = wb2.active
        headers2 = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]

        units_created = 0
        links_created = 0

        for row in ws2.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers2, row))
            if not data.get("IDPLOCHY"):
                continue

            code = str(data.get("IDPLOCHY") or "").strip()
            idk = int(data.get("KARTY.IDK") or 0)
            card = cards_map.get(idk)

            if not card:
                self.stdout.write(f"  Karta nenalezena pro IDK={idk}, plocha: {code}")
                continue

            m2 = data.get("m2")
            rate = data.get("SazbaKcMRok")

            unit, created = Unit.objects.get_or_create(
                code=code,
                site=site,
                defaults={
                    "name": code,
                    "purpose": str(data.get("Ucel") or "").strip(),
                    "area_m2": float(m2) if m2 else None,
                    "rate_per_m2_year": float(rate) if rate else None,
                    "unit_type": str(data.get("Jednotka") or "m2").strip(),
                }
            )
            if created:
                units_created += 1

            from core.models import CardUnit
            if not CardUnit.objects.filter(card=card, unit=unit).exists():
                CardUnit.objects.create(card=card, unit=unit)
                links_created += 1

        self.stdout.write(f"Ploch vytvořeno: {units_created}")
        self.stdout.write(f"Propojení vytvořeno: {links_created}")
        self.stdout.write(self.style.SUCCESS("\nImport dokončen!"))
