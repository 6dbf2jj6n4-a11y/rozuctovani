"""
Import klicu pro areal NJ z JEDNOHO spojeneho Excel souboru (na rozdil
od FM, kde jsou 4 zvlastni soubory Klice_Elektro/Voda/Teplo/Ostatni.xlsx).
Radky se rozlisuji sloupcem "kategorie" (ELEKTRO / VODA / TEPLO / OSTATNÍ),
sloupec s kodem meridla/OM se vzdy jmenuje "EL_KodOM" bez ohledu na
kategorii (dano zdrojovym souborem).

Pro kazdou kategorii se pouzije stejna logika jako u prislusneho
FM-specifickeho prikazu:

  ELEKTRO/VODA/TEPLO (merene, meridlo se hleda podle kodu v Meter.code):
    K_CELKU  -> vaha (hodnota = sloupec Jednotek, NE Podil - viz
                import_klice_teplo.py pro zduvodneni). Typ klice je
                weighted_count, nebo person_count pokud kod odpovida
                TUV (teplá užitková voda).
    K_PLOSE  -> fixed_amount (hodnota = Mplochy * KCzaM / 12; pokud IDPLOCHY
                oznacuje jednu konkretni mistnost, dohleda se i Unit a ulozi
                se do AllocationKey.unit - jen informativni)
    PEVNA_KC -> podle Jednotek a PevnaKC:
                Jednotek == 0 -> klic se nevytvari (neuctuje se)
                Jednotek != 0, PevnaKC == 0 -> weighted_count, hodnota=Jednotek
                Jednotek != 0, PevnaKC != 0 -> fixed_amount, hodnota=PevnaKC/12
                (POZOR: u NJ je PevnaKC rocni castka, na rozdil od FM
                Klice_*.xlsx, kde uz je to mesicni - potvrzeno uzivatelem),
                deduct_from_pool=False (pausal nezavisly na hlavnim odberu)

  OSTATNÍ (nemerene - ostraha, uklid, svoz odpadu...): OM kod se
    namapuje na nazev polozky pres Zasobnik_sluzeb.xlsx (stejne jako
    import_klice_ostatni.py), ServicePoolItem se hleda podle
    (site, name, meter=None).

Pouziti:
  python manage.py import_klice_nj cesta/k/klice_NJ.xlsx --site NJ
  (volitelne --zasobnik cesta/k/Zasobnik_sluzeb.xlsx, jinak se hleda
  ve stejne slozce jako tento management command)
"""
from pathlib import Path

import openpyxl
from decimal import Decimal, ROUND_HALF_UP
from django.core.management.base import BaseCommand
from core.models import Site, Meter, ServicePoolItem, ClientCard, AllocationKey, Unit


DEFAULT_ZASOBNIK_NAME = "Zasobnik_sluzeb.xlsx"

METERED_KATEGORIE = {"ELEKTRO", "VODA", "TEPLO"}


class Command(BaseCommand):
    help = "Importuje klice pro NJ (vsechny kategorie v jednom souboru, sloupec 'kategorie')"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--site", type=str, default="NJ")
        parser.add_argument(
            "--zasobnik", type=str, default=None,
            help="Cesta k Zasobnik_sluzeb.xlsx (pro OSTATNI radky). Vychozi: "
                 "stejna slozka jako tento command.",
        )

    def load_om_to_name(self, zasobnik_path, site_code):
        """Vrati dict {OM_kod: Popis} pro radky tridy OSTATNI a dany areal."""
        wb = openpyxl.load_workbook(zasobnik_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}

        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["class"]] != "OSTATNÍ":
                continue
            areal = str(row[idx["Areal"]] or "").strip()
            if areal.upper() != site_code.upper():
                continue
            om = str(row[idx["OM"]] or "").strip()
            popis = str(row[idx["Popis"]] or "").strip()
            if om:
                mapping[om] = popis
        return mapping

    def handle(self, *args, **options):
        site = Site.objects.filter(name__icontains=options["site"]).first()
        if not site:
            self.stdout.write(self.style.ERROR(f"Areál '{options['site']}' nenalezen."))
            return

        zasobnik_path = options["zasobnik"] or (Path(__file__).resolve().parent / DEFAULT_ZASOBNIK_NAME)
        if not Path(zasobnik_path).exists():
            self.stdout.write(self.style.ERROR(
                f"Soubor se zásobníkem nenalezen: {zasobnik_path} "
                f"(zadej --zasobnik cesta/k/Zasobnik_sluzeb.xlsx)"
            ))
            return
        om_to_name = self.load_om_to_name(zasobnik_path, options["site"])

        wb = openpyxl.load_workbook(options["xlsx_path"], read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        created = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            kategorie = str(data.get("kategorie") or "").strip().upper()
            card_desc = str(data.get("PopisKarty") or "").strip()
            aktivni = str(data.get("Aktivni") or "").strip().lower()
            om_code = str(data.get("EL_KodOM") or "").strip()
            typ = str(data.get("TYP_Polozky") or "").strip()
            mplochy = data.get("Mplochy")
            kczam = data.get("KCzaM")
            pevna_kc = data.get("PevnaKC")
            jednotek = data.get("Jednotek")
            idplochy = str(data.get("IDPLOCHY") or "").strip()

            if not card_desc or not om_code or aktivni != "zapnuto":
                skipped += 1
                continue

            card = ClientCard.objects.filter(description=card_desc, is_active=True).first()
            if not card:
                self.stdout.write(f"  Karta nenalezena: {card_desc}")
                skipped += 1
                continue

            # --- dohledani ServicePoolItem podle kategorie ---
            if kategorie in METERED_KATEGORIE:
                meter = Meter.objects.filter(code=om_code, site=site).first()
                if not meter:
                    self.stdout.write(f"  Měřidlo nenalezeno: {om_code} ({kategorie})")
                    skipped += 1
                    continue
                service_item = ServicePoolItem.objects.filter(meter=meter, site=site).first()
                if not service_item:
                    root = meter
                    while root.parent_meter:
                        root = root.parent_meter
                    service_item = ServicePoolItem.objects.filter(meter=root, site=site).first()
                if not service_item:
                    self.stdout.write(f"  ServicePoolItem nenalezen pro měřidlo: {om_code}")
                    skipped += 1
                    continue
            elif kategorie == "OSTATNÍ":
                meter = None
                service_name = om_to_name.get(om_code)
                if not service_name:
                    self.stdout.write(f"  OM kód nenalezen v zásobníku ({options['site']}): {om_code}")
                    skipped += 1
                    continue
                service_item = ServicePoolItem.objects.filter(
                    site=site, name=service_name, meter__isnull=True
                ).first()
                if not service_item:
                    self.stdout.write(
                        f"  ServicePoolItem nenalezen: '{service_name}' (OM {om_code}) - "
                        f"zkontroluj, jestli je Zásobník naimportovaný z aktuálního Zasobnik_sluzeb.xlsx."
                    )
                    skipped += 1
                    continue
            else:
                self.stdout.write(f"  Neznámá kategorie: '{kategorie}' ({card_desc} / {om_code})")
                skipped += 1
                continue

            # --- vypocet hodnoty a typu klice ---
            deduct_from_pool = True
            unit = None
            if typ == "K_CELKU":
                jednotek_val = Decimal(str(jednotek)) if jednotek not in (None, "") else None
                if not jednotek_val:
                    self.stdout.write(f"  Přeskočeno (Jednotek=0/chybí): {card_desc} / {om_code}")
                    skipped += 1
                    continue
                value = jednotek_val.quantize(Decimal("0.0001"))
                allocation_type = "person_count" if "TUV" in om_code.upper() else "weighted_count"
            elif typ == "K_PLOSE":
                allocation_type = "fixed_amount"
                if mplochy and kczam and float(kczam) > 0:
                    value = (Decimal(str(mplochy)) * Decimal(str(kczam)) / 12).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                    # IDPLOCHY oznacuje konkretni plochu jen pokud je to jedna
                    # mistnost - u souctove castky pres vice ploch byva 0/prazdne
                    # a plocha se nedoplni.
                    if idplochy and idplochy != "0":
                        unit = Unit.objects.filter(site=site, name=idplochy).first()
                else:
                    value = None
            elif typ == "PEVNA_KC":
                jednotek_val = Decimal(str(jednotek)) if jednotek not in (None, "") else Decimal("0")
                pevna_kc_val = Decimal(str(pevna_kc)) if pevna_kc not in (None, "") else Decimal("0")
                if jednotek_val == 0:
                    self.stdout.write(
                        f"  Přeskočeno (Jednotek=0, neúčtuje se): {card_desc} / {om_code}"
                    )
                    skipped += 1
                    continue
                elif pevna_kc_val == 0:
                    allocation_type = "weighted_count"
                    value = jednotek_val.quantize(Decimal("0.0001"))
                else:
                    # U NJ je PevnaKC u pevnych castek ROCNI cena (potvrzeno uzivatelem,
                    # na rozdil od FM Klice_*.xlsx, kde uz je to mesicni castka primo)
                    allocation_type = "fixed_amount"
                    value = (pevna_kc_val / 12).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    deduct_from_pool = False
                    self.stdout.write(
                        f"  + pausal {card_desc} / {om_code}: {pevna_kc_val} Kč/rok -> {value} Kč/měsíc"
                    )
            else:
                self.stdout.write(f"  Neznámý typ: {typ} ({card_desc} / {om_code})")
                skipped += 1
                continue

            key, was_created = AllocationKey.objects.update_or_create(
                client_card=card,
                service_item=service_item,
                meter=meter if (kategorie in METERED_KATEGORIE and typ == "K_CELKU") else None,
                defaults={
                    "allocation_type": allocation_type,
                    "value": value,
                    "deduct_from_pool": deduct_from_pool,
                    "unit": unit,
                },
            )

            if was_created:
                created += 1
                self.stdout.write(f"  + {card_desc} / {om_code} ({kategorie}, {allocation_type}): {value}")
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nHotovo: {created} vytvořeno, {updated} aktualizováno, {skipped} přeskočeno."
        ))
