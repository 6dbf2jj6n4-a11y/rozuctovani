"""
Diagnostika (jen cteni, nic nemeni): porovna typ a hodnotu klicu v DB proti
tomu, co by mel byt podle zdrojoveho klice_NJ.xlsx a logiky
import_klice_nj.py, pro merene kategorie (ELEKTRO/VODA/TEPLO) a radky
typu K_CELKU.

Duvod: u klienta Aktiv Novostav ma klic pro "hlavní odběr elektro NJ"
(meridlo E_AB1_10) v DB typ "Podružné měřidlo" (submeter), ale zdrojovy
soubor rika K_CELKU -> podle soucasne logiky import_klice_nj.py by mel byt
typu "weighted_count" (Podle vahy). U typu submeter se ale hodnota klice
(key.value) ve skutecnem vypoctu vubec nepouziva (podil = spotreba
podruzneho meridla / celkova spotreba hlavniho meridla) - misto vaseneho
podilu na zbytku se tak klientovi pocita cisty pomer dvou meridel, coz
dava uplne jiny (typicky mnohem mensi) vysledek. Tento skript zjisti,
kolika klicu/klientu se to tyka.

Pouziti:
  python manage.py diag_typy_klicu_nj
  (volitelne --xlsx cesta/k/klice_NJ.xlsx, jinak se pouzije kopie
  ulozena primo v repu: core/management/commands/klice_NJ.xlsx)
"""
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from core.models import AllocationKey, ClientCard, Meter, ServicePoolItem

METERED_KATEGORIE = {"ELEKTRO", "VODA", "TEPLO"}
DEFAULT_XLSX_NAME = "klice_NJ.xlsx"


class Command(BaseCommand):
    help = "Porovná typ/hodnotu klíčů NJ (ELEKTRO/VODA/TEPLO, K_CELKU) v DB proti zdrojovému Excelu."

    def add_arguments(self, parser):
        parser.add_argument("--xlsx", default=None)

    def handle(self, *args, **options):
        xlsx_path = options["xlsx"] or (Path(__file__).resolve().parent / DEFAULT_XLSX_NAME)
        if not Path(xlsx_path).exists():
            self.stdout.write(self.style.ERROR(f"Soubor nenalezen: {xlsx_path}"))
            return

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        mismatches = []
        checked = 0
        not_found = 0
        counts_by_kategorie = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            kategorie = str(data.get("kategorie") or "").strip().upper()
            if kategorie not in METERED_KATEGORIE:
                continue
            typ = str(data.get("TYP_Polozky") or "").strip()
            if typ != "K_CELKU":
                continue
            aktivni = str(data.get("Aktivni") or "").strip().lower()
            if aktivni != "zapnuto":
                continue

            card_desc = str(data.get("PopisKarty") or "").strip()
            om_code = str(data.get("EL_KodOM") or "").strip()
            jednotek = data.get("Jednotek")
            if jednotek in (None, "") or Decimal(str(jednotek)) == 0:
                continue

            expected_value = Decimal(str(jednotek)).quantize(Decimal("0.0001"))
            expected_type = "person_count" if "TUV" in om_code.upper() else "weighted_count"

            card = ClientCard.objects.filter(description=card_desc, is_active=True).first()
            meter = Meter.objects.filter(code=om_code).first()
            if card is None or meter is None:
                not_found += 1
                continue

            key = AllocationKey.objects.filter(client_card=card, meter=meter).first()
            checked += 1
            counts_by_kategorie[kategorie] = counts_by_kategorie.get(kategorie, 0) + 1

            if key is None:
                mismatches.append((kategorie, card_desc, om_code, "CHYBÍ KLÍČ V DB", None, None, expected_type, expected_value))
                continue

            if key.allocation_type != expected_type or key.value != expected_value:
                mismatches.append((
                    kategorie, card_desc, om_code, "NESEDÍ",
                    key.get_allocation_type_display(), key.value,
                    expected_type, expected_value,
                ))

        self.stdout.write(self.style.WARNING(
            f"Zkontrolováno {checked} K_CELKU řádků (ELEKTRO/VODA/TEPLO), "
            f"{not_found} nedohledáno (karta/měřidlo), po kategoriích: {counts_by_kategorie}\n"
        ))

        if not mismatches:
            self.stdout.write(self.style.SUCCESS("Žádné neshody nenalezeny."))
            return

        self.stdout.write(self.style.ERROR(f"Nalezeno {len(mismatches)} neshod:\n"))
        for kategorie, card_desc, om_code, problem, actual_type, actual_value, expected_type, expected_value in mismatches:
            self.stdout.write(
                f"  [{kategorie}] {card_desc} / {om_code}: {problem} - "
                f"DB má typ={actual_type!r} hodnota={actual_value}, "
                f"Excel/import by dal typ={expected_type!r} hodnota={expected_value}"
            )
