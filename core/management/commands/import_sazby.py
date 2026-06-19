"""
Doplni cenu (Kc/m2/rok) do existujicich CardUnit zaznamu
z Excel souboru plochy.xlsx (sloupec SazbaKcMRok).
"""
import openpyxl
from django.core.management.base import BaseCommand
from core.models import CardUnit, Unit, ClientCard


class Command(BaseCommand):
    help = "Doplni sazbu Kc/m2/rok do CardUnit z Excel souboru"

    def add_arguments(self, parser):
        parser.add_argument("plochy_xlsx", type=str, help="Cesta k souboru plochy.xlsx")

    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(options["plochy_xlsx"], read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if not data.get("IDPLOCHY"):
                continue

            code = str(data.get("IDPLOCHY") or "").strip()
            idk = int(data.get("KARTY.IDK") or 0)
            rate = data.get("SazbaKcMRok")

            if not rate:
                skipped += 1
                continue

            card = ClientCard.objects.filter(external_id=idk).first()
            unit = Unit.objects.filter(code=code).first()

            if not card or not unit:
                self.stdout.write(f"  Nenalezeno: karta IDK={idk} nebo plocha {code}")
                skipped += 1
                continue

            card_unit = CardUnit.objects.filter(card=card, unit=unit).first()
            if not card_unit:
                self.stdout.write(f"  CardUnit nenalezen: {card} - {unit}")
                skipped += 1
                continue

            card_unit.rate_per_m2 = float(rate)
            card_unit.save()
            updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"Hotovo: {updated} sazeb doplněno, {skipped} přeskočeno."
        ))
