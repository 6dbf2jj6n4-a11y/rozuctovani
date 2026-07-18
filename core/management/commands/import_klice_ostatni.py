"""
Import klicu ostatni sluzby z Excel souboru Klice_Ostatni.xlsx.
Pro kazdy radek vytvori AllocationKey na prislusne karte klienta.

Mapovani TYP_Polozky:
  K_CELKU  -> vaha (hodnota = sloupec Jednotek, NE Podil - Podil je v
              puvodnim Excelu jen dopocitany napovedny udaj = Jednotek /
              soucet Jednotek pres vsechny karty se stejnym OSTATNI_KodOM;
              pouzitim primo Jednotek si system podily dopocitava sam a
              nezastarava pri zmene poctu najemcu). Typ klice je vzdy
              weighted_count (u OSTATNI polozek zatim neni znamy kod
              odpovidajici TUV).
  K_PLOSE  -> fixed_amount  (hodnota = Mplochy * KCzaM / 12 = mesicni pausal;
              pokud IDPLOCHY oznacuje jednu konkretni mistnost, dohleda se
              i Unit a ulozi se do AllocationKey.unit - jen informativni)
  PEVNA_KC -> podle sloupce Jednotek a PevnaKC:
              Jednotek == 0            -> klic se NEVYTVARI (polozka se aktualne
                                           neuctuje, napr. INTERNET u nekterych karet)
              Jednotek != 0, PevnaKC == 0 -> weighted_count, hodnota = Jednotek
                                           (Jednotek je vaha/pocet kusu pro rozpocet
                                           nakladu zadaneho rucne, napr. pocet hasicich
                                           pristroju u revize hasicich pristroju)
              Jednotek != 0, PevnaKC != 0 -> fixed_amount, hodnota = PevnaKC / 12
                                           (PevnaKC je ROCNI castka - potvrzeno
                                           uzivatelem), deduct_from_pool=False

Polozky tridy OSTATNI v Zasobniku NEMAJI mericí (nejsou merene - ostraha,
uklid, svoz odpadu apod.), takze na rozdil od Elektro/Voda/Teplo tu nejde
hledat ServicePoolItem pres Meter. Misto toho se OSTATNI_KodOM (napr.
OSTRAHA, INTERNET, UKLID_A) namapuje na nazev polozky (sloupec "Popis")
pomoci zdrojoveho souboru Zasobnik_sluzeb.xlsx (sloupce OM / Popis / Areal
/ class), ktery je stejny soubor jako pri importu import_zasobnik_sluzeb_v2,
a ServicePoolItem se pak hleda podle (site, name, meter=None).

Pouziti:
  python manage.py import_klice_ostatni cesta/k/Klice_Ostatni.xlsx --site FM
  (volitelne --zasobnik cesta/k/Zasobnik_sluzeb.xlsx, jinak se hleda
  ve stejne slozce jako tento management command)
"""
from pathlib import Path

import openpyxl
from decimal import Decimal, ROUND_HALF_UP
from django.core.management.base import BaseCommand
from core.models import Site, ServicePoolItem, ClientCard, AllocationKey, Unit


TYP_TO_ALLOCATION = {
    "K_CELKU": "weighted_count",
    "K_PLOSE": "fixed_amount",
    "PEVNA_KC": "fixed_amount",
}

DEFAULT_ZASOBNIK_NAME = "Zasobnik_sluzeb.xlsx"


class Command(BaseCommand):
    help = "Importuje klice ostatni sluzby (AllocationKey) z Excel souboru"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--site", type=str, default="FM")
        parser.add_argument(
            "--zasobnik", type=str, default=None,
            help="Cesta k Zasobnik_sluzeb.xlsx (pro dohledani nazvu polozky "
                 "podle OM kodu). Vychozi: stejna slozka jako tento command.",
        )

    def load_om_to_name(self, zasobnik_path, site_code):
        """Vrati dict {OM_kod: Popis} pro radky tridy OSTATNI a dany areal."""
        wb = openpyxl.load_workbook(zasobnik_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers)}

        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["class"]] != "OSTATNÍ":
                continue
            areal = str(row[idx["Areal"]] or "").strip()
            if areal.upper() != site_code.upper():
                continue
            om = str(row[idx["OM"]] or "").strip()
            popis = str(row[idx["Popis"]] or "").strip()
            if om:
                mapping[om] = popis
        return mapping

    def handle(self, *args, **options):
        site = Site.objects.filter(name__icontains=options["site"]).first()
        if not site:
            self.stdout.write(self.style.ERROR(f"Areál '{options['site']}' nenalezen."))
            return

        zasobnik_path = options["zasobnik"] or (Path(__file__).resolve().parent / DEFAULT_ZASOBNIK_NAME)
        if not Path(zasobnik_path).exists():
            self.stdout.write(self.style.ERROR(
                f"Soubor se zásobníkem nenalezen: {zasobnik_path} "
                f"(zadej --zasobnik cesta/k/Zasobnik_sluzeb.xlsx)"
            ))
            return
        om_to_name = self.load_om_to_name(zasobnik_path, options["site"])

        wb = openpyxl.load_workbook(options["xlsx_path"], read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        created = 0
        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            card_desc = str(data.get("PopisKarty") or "").strip()
            aktivni = str(data.get("Aktivni") or "").strip().lower()
            om_code = str(data.get("OSTATNI_KodOM") or "").strip()
            typ = str(data.get("TYP_Polozky") or "").strip()
            mplochy = data.get("Mplochy")
            kczam = data.get("KCzaM")
            pevna_kc = data.get("PevnaKC")
            jednotek = data.get("Jednotek")
            idplochy = str(data.get("IDPLOCHY") or "").strip()

            if not card_desc or not om_code or aktivni != "zapnuto":
                skipped += 1
                continue

            allocation_type = TYP_TO_ALLOCATION.get(typ)
            if not allocation_type:
                self.stdout.write(f"  Neznámý typ: {typ} ({card_desc} / {om_code})")
                skipped += 1
                continue

            card = ClientCard.objects.filter(description=card_desc, is_active=True).first()
            if not card:
                self.stdout.write(f"  Karta nenalezena: {card_desc}")
                skipped += 1
                continue

            service_name = om_to_name.get(om_code)
            if not service_name:
                self.stdout.write(f"  OM kód nenalezen v zásobníku ({options['site']}): {om_code}")
                skipped += 1
                continue

            service_item = ServicePoolItem.objects.filter(
                site=site, name=service_name, meter__isnull=True
            ).first()
            if not service_item:
                self.stdout.write(
                    f"  ServicePoolItem nenalezen: '{service_name}' (OM {om_code}) - "
                    f"zkontroluj, jestli je Zásobník naimportovaný z aktuálního Zasobnik_sluzeb.xlsx."
                )
                skipped += 1
                continue

            # Vypocet hodnoty klice
            deduct_from_pool = True
            unit = None
            if typ == "K_CELKU":
                jednotek_val = Decimal(str(jednotek)) if jednotek not in (None, "") else None
                if not jednotek_val:
                    self.stdout.write(f"  Přeskočeno (Jednotek=0/chybí): {card_desc} / {service_name}")
                    skipped += 1
                    continue
                value = jednotek_val.quantize(Decimal("0.0001"))
            elif typ == "K_PLOSE":
                if mplochy and kczam and float(kczam) > 0:
                    # mesicni pausal = plocha * cena/m2/rok / 12
                    value = (Decimal(str(mplochy)) * Decimal(str(kczam)) / 12).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                    # IDPLOCHY oznacuje konkretni plochu jen pokud je to jedna
                    # mistnost - u souctove castky pres vice ploch byva 0/prazdne
                    # a plocha se nedoplni.
                    if idplochy and idplochy != "0":
                        unit = Unit.objects.filter(site=site, name=idplochy).first()
                else:
                    value = None
            elif typ == "PEVNA_KC":
                jednotek_val = Decimal(str(jednotek)) if jednotek not in (None, "") else Decimal("0")
                pevna_kc_val = Decimal(str(pevna_kc)) if pevna_kc not in (None, "") else Decimal("0")
                if jednotek_val == 0:
                    # Jednotek=0 = polozka se u teto karty aktualne neuctuje
                    self.stdout.write(
                        f"  Přeskočeno (Jednotek=0, neúčtuje se): {card_desc} / {service_name}"
                    )
                    skipped += 1
                    continue
                elif pevna_kc_val == 0:
                    # castka se neplati primo - Jednotek je vaha pro rozpocet
                    # nakladu, ktery se zada rucne (CostEntry / vychozi castka)
                    allocation_type = "weighted_count"
                    value = jednotek_val.quantize(Decimal("0.0001"))
                else:
                    # PevnaKC je rocni castka
                    allocation_type = "fixed_amount"
                    value = (pevna_kc_val / 12).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    deduct_from_pool = False
            else:
                value = None

            key, was_created = AllocationKey.objects.update_or_create(
                client_card=card,
                service_item=service_item,
                meter=None,
                defaults={
                    "allocation_type": allocation_type,
                    "value": value,
                    "deduct_from_pool": deduct_from_pool,
                    "unit": unit,
                },
            )

            if was_created:
                created += 1
                self.stdout.write(f"  + {card_desc} / {service_name} ({allocation_type}): {value}")
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nHotovo: {created} vytvořeno, {updated} aktualizováno, {skipped} přeskočeno."
        ))
