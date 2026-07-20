"""
Import stavu meridel z Excel souboru Spotreby_2026_FM.xlsx.
Zpracovava listy ELEKTRO, VODA, TEPLO.
Pro kazde meridlo nacte stav ke dni "1. 6. 2026" a "1. 7. 2026"
a ulozi jako MeterReading pro obdobi 05/2026 a 06/2026.

Pouziti:
  python manage.py import_odecty cesta/k/souboru.xlsx --site FM --year 2026
"""
import openpyxl
from datetime import date
from django.core.management.base import BaseCommand
from core.models import Site, Meter, Period, MeterReading


SHEETS = ["ELEKTRO", "VODA", "TEPLO"]

# Mapovani nazvu sloupcu na mesice
MONTH_MAP = {
    "1. 1. 2026": (2026, 1),
    "1. 2. 2026": (2026, 2),
    "1. 3. 2026": (2026, 3),
    "1. 4. 2026": (2026, 4),
    "1. 5. 2026": (2026, 5),
    "1. 6. 2026": (2026, 6),
    "1. 7. 2026": (2026, 7),
}


class Command(BaseCommand):
    help = "Importuje stavy meridel z Excel souboru (listy ELEKTRO/VODA/TEPLO)"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument("--site", type=str, required=True)
        parser.add_argument(
            "--months",
            type=str,
            default="6,7",
            help="Mesice ke zpracovani jako stavy (vychozi: 6,7 = stav k 1.6 a 1.7 = spotreba za cerven)",
        )
        parser.add_argument("--year", type=int, default=2026)

    def handle(self, *args, **options):
        site = Site.objects.filter(name__icontains=options["site"]).first()
        if not site:
            self.stdout.write(self.style.ERROR(f"Areál '{options['site']}' nenalezen."))
            return

        months = [int(m) for m in options["months"].split(",")]
        year = options["year"]

        wb = openpyxl.load_workbook(options["xlsx_path"], read_only=True, data_only=True)

        created = 0
        updated = 0
        skipped = 0

        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = list(rows[0])
            # Najdeme indexy sloupcu pro pozadovane mesice
            col_indices = {}
            for i, h in enumerate(headers):
                h_str = str(h).strip() if h else ""
                if h_str in MONTH_MAP:
                    y, m = MONTH_MAP[h_str]
                    if y == year and m in months:
                        col_indices[m] = i

            if not col_indices:
                self.stdout.write(f"  {sheet_name}: nenalezeny sloupce pro rok {year}, měsíce {months}")
                continue

            self.stdout.write(f"\n--- {sheet_name} ---")

            for row in rows[1:]:
                code = str(row[0]).strip() if row[0] else ""
                if not code or code == "STAVY":
                    continue
                if code == "SPOTŘEBY":
                    # Pod tabulkou STAVY nasleduje samostatna tabulka SPOTREBY se
                    # stejnymi kody meridel, ale jinymi sloupci (Koef mista Jedn.,
                    # data od 2022) - kdybychom pokracovali dal, presly bychom
                    # tyhle radky se STEJNYMI indexy sloupcu jako STAVY tabulka a
                    # tise prepsali spravne odecty spatnymi cisly. Zbytek listu
                    # (SPOTREBY + souhrnove radky pod ni) se tedy preskoci cely.
                    break

                meter = Meter.objects.filter(site=site, code=code).first()
                if not meter:
                    skipped += 1
                    continue

                for month, col_idx in col_indices.items():
                    value = row[col_idx]
                    if value is None or value == 0:
                        continue
                    try:
                        value = float(value)
                    except (TypeError, ValueError):
                        continue

                    # Stav k 1.X.YYYY je odecet pro predchozi mesic (X-1)
                    # Stav k 1.6. je odecet pro 05/2026
                    # Stav k 1.7. je odecet pro 06/2026
                    period_month = month - 1 if month > 1 else 12
                    period_year = year if month > 1 else year - 1
                    reading_date = date(year, month, 1)

                    period, _ = Period.objects.get_or_create(
                        year=period_year,
                        month=period_month,
                        defaults={"status": "open"},
                    )

                    reading, was_created = MeterReading.objects.update_or_create(
                        meter=meter,
                        period=period,
                        defaults={
                            "value": value,
                            "reading_date": reading_date,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1

                    self.stdout.write(
                        f"  {code} / {period_month:02d}/{period_year}: {value}"
                    )

        self.stdout.write(self.style.SUCCESS(
            f"\nHotovo: {created} odečtů vytvořeno, {updated} aktualizováno, {skipped} měřidel nenalezeno."
        ))
